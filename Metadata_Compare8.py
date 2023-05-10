'''
Developer: Jithender Thota
Date: 02-Mar-2023
This script compares two metadata files and the differences identified between the two files are written to an excel file.
Metadata files can have either .app or .txt extension.
This script will give wierd results if both the files are not symmetric. Also the automation fails if the number of custom 
dimensions are not same in both the files. This automation works only for applications with 4 custom dimensions. Below two lines
must be modified if it has to work for an application with different number of custom dimensions.

sEnti = 'Label,DefaultValueID,AllowAdjustments,IsICP,AllowChildrenAdjs,SecurityClassID,UserDefined1,UserDefined2,UserDefined3,HoldingCompany,EAPSecurityClassID,DefaultParent,Descriptions'
sAcco = 'Label,AccountType,IsCalculated,IsConsolidated,IsICP,PlugAcct,Custom1TopMember,Custom2TopMember,Custom3TopMember,Custom4TopMember,NumDecimalPlaces,UsesLineItems,EnableCustom1Aggr,EnableCustom2Aggr,EnableCustom3Aggr,EnableCustom4Aggr,UserDefined1,UserDefined2,UserDefined3,XBRLTags,SecurityClass,ICPTopMember,EnableDataAudit,CalcAttribute,SubmissionGroup,DefaultParent,Descriptions'
	

This automation requires 3 parameters
1) Two metadata files to be comapred
2) Path to the above metadata files

if the above parameters are not provided while running the script then it will consider ABTPLNQA_Metadata.app metadata file 
as file1 and ABTPROD_Metadata.app is considered for file2. If no path is provided, then system will look for the metadata 
files in the current working directory.

setup
=====
Raise a request to install "Python Python 3.9.7" or "Anaconda Anaconda 5.3.1" from the snow portal.
Recommendation is "Anaconda Anaconda 5.3.1", because this comes up with many libraries that are not available in python 3.9.7
If python 3.9.7 is installed.. open command prompt and run the below command, this command will install the python library openpyxl
python -m pip install openpyxl


Command to run the script:
==========================
1) Browse to the directory where this script is saved using the cmd
2) Use either of the below commands to run the automation
    2.a) python Metadata_Compare.py 
         This will compare the files ABTPLNQA_Metadata.app and ABTPROD_Metadata.app
    2.b) python Metadata_Compare.py -f1 <Metadata file1> -f2 <Metadata file2> -p <Path to these metadata files>  
         provide the metadata filenames and path without using the angular braces, also save the two metadata files in the same location
         
3) python Metadata_Compare.py --help to get the help on this script. This will basically gives the usage of the script


Libraries 
=========
Only the important libraries usage is provided here

openpyxl: Used to write the results to an excel file. Results are actually written to a csv file, styling cannot be done
in a csv file. So the results are copied from CSV file to an excel file and then styled using this library.

difflib: compares two files and generates a single file with all the differences, this library is used for optimization purpose.
Before using this library automation was taking some 30 to 50 mins depending on the size of the metadata files. After using this
library it completes in seconds.

Algorithm:
=========
1) Directories are created to store temporary files created during the script execution.
2) Some metadata files has spaces and some may not have spaces. So spaces are removed first in the metadata files
3) Metadata files are then split into small metadata files by the sections in the metadata and then stored in their respective folders
4) Then these small metadata files are compared using the diff library and then results are written to the CSV file
5) Then the results are copied to an xls file to allow some formatting
6) All the temporary files are deleted


'''


import openpyxl
from openpyxl.styles import Font, Color, PatternFill
from openpyxl import Workbook
import difflib
import os
import sys
from datetime import datetime
import csv
import argparse
import re
from pathlib import Path
import shutil
import textwrap
import time

now = datetime.now()

def getFontStyle():
    return Font(size=14,bold=True)

def getBackGroundColor():
        return PatternFill(patternType='solid', fgColor='92D2E2')
        
def getCurrentTime():
    return str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

def printLines(textToPrint):
    print(20 * "*")
    print("INFO: {}: {}" .format(getCurrentTime(), textToPrint))
    
def printWarningLines(textToPrint):
    print(20 * "(#")
    print("WARNING: {}: {}" .format(getCurrentTime(), textToPrint))
    
def printLine(textToPrint):
    print("INFO: {}: {}" .format(getCurrentTime(), textToPrint))

def createResultsFile(differencesfilePath):
    
    global excelFilePath
    printLines("creating the Results file")
    excelFilePath = differencesfilePath.split(".")[0] + ".xlsx"
    #print("excelFilePath: " + excelFilePath)
    book =  Workbook()
    book.save(excelFilePath)
    printLine("completed ....")
    writeToResultsFile()
 

        
def writeToResultsFile():
    printLines("updating the results file with differences")
    wb = openpyxl.load_workbook(filename=excelFilePath)
    ws = wb['Sheet']
    rowNumber = 0
      
    with open(differencesfilePath, encoding="cp1252") as resultsCSVFile:
        for line in resultsCSVFile:
            #print(line)
            rowNumber = rowNumber + 1
            aLine = line.split(",")
            for columnNumber, item in enumerate(aLine):
                ws.cell(row = rowNumber, column=columnNumber+1).value = item
                if rowNumber == 1:
                    cellNumber = ws.cell(row = rowNumber, column=columnNumber+1)
                    cellNumber.font = getFontStyle()
                    cellNumber.fill = getBackGroundColor()

    wb.save(excelFilePath)
    printLine("completed ....")
    
'''
Creating the temporary folders to store the intermediate metadata files
'''

def getFullPath(folderPath, fileToJoin):
    return os.path.join(folderPath, fileToJoin)

def bPathExists(sPath):
    if os.path.exists(sPath):
        return True
    return False
    
    
def createTempFolders(filePath, metadataFile1, metadataFile2):
	printLines("creating temp folders to store the temporary files")
	metadataFilesFolder = getFullPath(filePath, "Dimension_files")
	if not bPathExists(metadataFilesFolder):
		os.mkdir(metadataFilesFolder)
	if not bPathExists(getFullPath(metadataFilesFolder, metadataFile1)):
		os.mkdir(getFullPath(metadataFilesFolder, metadataFile1))
	if not bPathExists(getFullPath(metadataFilesFolder, metadataFile2)):
		os.mkdir(getFullPath(metadataFilesFolder, metadataFile2))
	if not bPathExists(os.path.join(metadataFilesFolder, "differences")):
		os.mkdir(getFullPath(metadataFilesFolder,"differences"))
	return metadataFilesFolder

def numberOfCustomDimensions(filePath,metadataFiles):
	anumOfCustDimensions = []
	printLines("Validating the number of custom dimensions in the given metadata files")
	for file in metadataFiles:
		with open(getFullPath(filePath,file), 'r', encoding="cp1252", newline="") as metadataFile:
			for line in metadataFile:
				if "!CUSTOM_ORDER" in line:
					customDimensions = line.split("=")[1]
					numberofCustDimensions = len(customDimensions.split(";"))
					anumOfCustDimensions.append(numberofCustDimensions)

					break
	
	if anumOfCustDimensions[0] != anumOfCustDimensions[1]:
		raise Exception("Two files are not having symmetric number of custom dimensions...\
						 terminating the process")
	else:
		printLine("Files are validated .. comparison will begin")
		return customDimensions
				
'''
Some metadata files have spaces and few metadata files might not have spaces. 
So removing the spaces in the both the metadata files before running the comparison

'''
def trimMetadataFiles(filePath, metadataFiles):
	for file in metadataFiles:
		source_file = open(getFullPath(filePath,file), encoding="cp1252")
		destination_file = open(getFullPath(filePath,file.split(".")[0] + "_v1.txt"), "w", encoding="cp1252")
		for line in source_file:
			line_split = line.split(";")
			new_line = ""
			for i in range(0,len(line_split)):
				spaces_removed_line_split = line_split[i].strip()
				if i==0:
					new_line = spaces_removed_line_split
				else:
					new_line = new_line + ";" + spaces_removed_line_split
			destination_file.write(new_line + "\n")
		source_file.close()
		destination_file.close()

'''
For quick comparision Metadata is split into different fies based on members and hierarchy, for ex: Account dimension has two sections
in the metadata file, Members section and Hierarchy section. So account dimension is split into two different files, one file containing
only members section and the other file containing Hierarchy section.
'''
		
def splitFile(filePath, metadataFiles):
	aDimensionsInBothFiles = []    
	for file in metadataFiles:
		processOrder = 0
		aDimensionsInEachFile = []		
		printLines("Separating members and hierarchies sections in metadata file" + file + " to separate files")
		trimmedFile = getFullPath(filePath, file.split(".")[0]) + "_v1.txt"
		destinationPath = getFullPath(filePath,"Dimension_files")
		destinationPath = getFullPath(destinationPath,file.split(".")[0])
		with open(trimmedFile, encoding="cp1252") as file:
			dimensionArray = ['!APPLICATION_SETTINGS', '!CURRENCIES', '!MEMBERS','!HIERARCHIES', '!CONSOLIDATION_METHODS']
			nonDimensionArray = ['!FILE_FORMAT', '!VERSION', '!CUSTOM_ORDER', '!LABEL']
			fileName = ""
			for line in file:
				if not line.strip() == '':
					aline = line.split("=")
					if aline[0] not in nonDimensionArray:
						if line[0] == "!" and (aline[0].strip().upper() in dimensionArray):
							#processOrder = processOrder + 1
							if len(aline) == 1:
								#fileName = str(processOrder) + "." + aline[0][1:].strip().upper()
								fileName = aline[0][1:].strip().upper()
							else:
								#fileName = str(processOrder) + "." + aline[1].strip() + aline[0][1]
								fileName = aline[1].strip() + aline[0][1]
							aDimensionsInEachFile.append(fileName)
							output_file = open(os.path.join(destinationPath, fileName +".txt"), "w", encoding="cp1252")
													
						else:
							output_file.write(line)
		aDimensionsInBothFiles.append(aDimensionsInEachFile)
		printLine("Completed ....")
		#print("aDimensionsInEachFile")
		#printLine(aDimensionsInEachFile)
	#print("aDimensionsInBothFiles")
	#print(aDimensionsInBothFiles)
	return aDimensionsInBothFiles
		
'''
difflib library is used to identify the differences in two files, differences identified are written to the file diff.txt.
This diff.txt has differences from both the metadata files. These differences are split into two different files.
'''
def processBlock(metadataFile):
	anonStandardDimension = ["APPLICATION_SETTINGS", "CURRENCIES", "CONSOLIDATION_METHODS"]
	#print(metadataFile)
	if metadataFile.split(".")[0] in anonStandardDimension:
		processingDimension = metadataFile.split(".")[0]
	elif metadataFile.split(".")[0][-1] == "H":
		processingDimension = metadataFile.split(".")[0][:-1] + " Hierarchies"
	elif metadataFile.split(".")[0][-1] == "M":
		processingDimension = metadataFile.split(".")[0][:-1] + " Members"
	return processingDimension
	
def findDifferences(differences_FilePath, metadataFolder1, metadataFolder2, metadataFile):
	printLines("Finding discrepancies in " + processBlock(metadataFile))
	fileFromFolder1 = getFullPath(os.path.join(differences_FilePath,metadataFolder1), metadataFile)
	fileFromFolder2 = getFullPath(os.path.join(differences_FilePath,metadataFolder2), metadataFile)
	with open(fileFromFolder1, "r", encoding="cp1252") as f1:
		with open(fileFromFolder2, "r", encoding="cp1252") as f2:
			diff = difflib.unified_diff(f1.readlines(), f2.readlines(),fromfile='f1', tofile='f2',)
			with open(getFullPath(differences_FilePath,"differences/diff.txt"),"w", encoding="cp1252") as diff_file:
				for line in diff:
					diff_file.write(line)
	printLine("Completed ....")
	splitDifferences(differences_FilePath, metadataFile.split(".")[0])
		
def splitDifferences(filePath, dimension):
	file1 = open(getFullPath(filePath,"differences/file1.txt"),"w", encoding="cp1252")
	file2 = open(getFullPath(filePath,"differences/file2.txt"),"w", encoding="cp1252")
	
	with open(getFullPath(filePath,"differences/diff.txt"), "r", encoding="cp1252") as difference:
		for line in difference:
			if line[0] == "-" and len(line)>2 and line[1] != "-":
				file1.write(line[1:])
			if line[0] == "+" and len(line)>2 and line[1] != "+":
				file2.write(line[1:])
	file1.close()
	file2.close()
	#print(listOfCustDimensions)
	if not os.stat(getFullPath(filePath,"differences/file1.txt")).st_size == 0 or not os.stat(getFullPath(filePath,"differences/file2.txt")).st_size == 0:
		compare_files(getFullPath(filePath, "differences"), dimension)
		
'''
This is function which actually compares two metadata files and writes the results to the results file
'''
def getMetadataSection(dimension):
    metadataProperty = ""
    if dimension[-1].upper() == "H":
        metadataProperty = "Hierarchy"
    elif dimension[-1].upper() == "M":
        metadataProperty = "Member"
    return metadataProperty

def compare_files(diffFilePath, dimension):
	sCurr = 'Label,Scale,TranslationOperator,DisplayInICT,Descriptions'
	sScen = 'Label,DefaultFreq,DefaultView,ZeroViewForNonadj,ZeroViewForAdj,ConsolidateYTD,UserDefined1,UserDefined2,UserDefined3,SupportsProcessManagement,SecurityClass,MaximumReviewLevel,UsesLineItems,EnableDataAudit,DefFreqForICTrans,PhasedSubStartYear,DefaultParent,Descriptions'
	sEnti = 'Label,DefaultValueID,AllowAdjustments,IsICP,AllowChildrenAdjs,SecurityClassID,UserDefined1,UserDefined2,UserDefined3,HoldingCompany,EAPSecurityClassID,DefaultParent,Descriptions'
	#sAcco = 'Label, AccountType, IsCalculated, IsConsolidated, IsICP, PlugAcct, NumDecimalPlaces, usesLineItems, UserDefined1,UserDefined2, USerDefined3, XBRLTags, SecurityClass, ICPTopMember, EnableDataAudit, CalcAttribute, SubmissionGroup, DefaultParent, Descriptions'
	sAcco = 'Label,AccountType,IsCalculated,IsConsolidated,IsICP,PlugAcct,Custom1TopMember,Custom2TopMember,Custom3TopMember,Custom4TopMember,NumDecimalPlaces,UsesLineItems,EnableCustom1Aggr,EnableCustom2Aggr,EnableCustom3Aggr,EnableCustom4Aggr,UserDefined1,UserDefined2,UserDefined3,XBRLTags,SecurityClass,ICPTopMember,EnableDataAudit,CalcAttribute,SubmissionGroup,DefaultParent,Descriptions'
	
	sCust = 'Label,IsCalculated,SwitchSignForFlow,SwitchTypeForFlow,UserDefined1,UserDefined2,UserDefined3,SecurityClass,SubmissionGroup,DefaultParent,Descriptions'
	sCons = 'Label,UsedByCalcRoutine,IsHoldingMethod,ToPercentControlComp,ToPercentControl,percentConsol,Control,Descriptions'
	aCurr = sCurr.split(",")
	aScen = sScen.split(",")
	aEnti = sEnti.split(",")
	aAcco = sAcco.split(",")
	aCust = sCust.split(",")
	aCons = sCons.split(",")
	aMemLength = [len(aCurr), len(aScen), len(aEnti),len(aAcco),len(aCust),len(aCons)]
	aMemArray = [aCurr,aScen, aEnti,aAcco,aCust,aCons]
	file1 = getFullPath(diffFilePath, "file1.txt")
	file2 = getFullPath(diffFilePath, "file2.txt")
	
	with open(differencesfilePath, "a", encoding="cp1252", newline="") as outputCSVFile:
		CSVWriter = csv.writer(outputCSVFile)
		for f in [file1, file2]:
			processfile12 = False
			processfile21 = False
			if f == file1:
				s_file = file1
				d_file = file2
				processfile12 = True
			else:
				s_file = file2
				d_file = file1
				processfile21 = True
			from_file = open(s_file, "r", encoding="cp1252")
			to_file = open(d_file, "r", encoding="cp1252")
			i = 0
			for line1 in from_file:
				to_file.seek(0,0)
				aline1 = line1.split(";")
				index = 0
				writtentoFile = False
				exists = False
				for item in aMemLength:
					if len(aline1) == item:
						aPropertyArray = aMemArray[index]
						break
					index = index + 1	
				
				if len(aline1) == 1:
					aline1Split = aline1[0].split("=")
					for line2 in to_file:
						aline2 = line2.split(";")
						if len(aline2) == 1:
							aline2Split = aline2[0].split("=")
							if aline1Split[0] == aline2Split[0]:
								exists = True
								writtentoFile = True
								if aline1Split[1] != aline2Split[1] and processfile12 == True:
									CSVWriter.writerow([dimension[:len(dimension)-1],aline1Split[0], "Value", aline1Split[1].strip(),aline2Split[1].strip()])
									break
					if exists == False and processfile12 == True:
						CSVWriter.writerow([dimension[:len(dimension)-1], line1.strip(), getMetadataSection(dimension),"", "Missing"])
					elif exists == False and processfile21 == True:
						CSVWriter.writerow([dimension[:len(dimension)-1], line1.strip(), getMetadataSection(dimension), "Missing"])
				elif len(aline1) == 2:
					for line2 in to_file:
						if line1.upper() == line2.upper():
							exists = True
							break
					if exists == False and processfile12 == True:
						CSVWriter.writerow([dimension[:len(dimension)-1], line1.strip(), getMetadataSection(dimension), "", "Missing"])
						writtentoFile = True
					elif exists == False and processfile21 == True:
						CSVWriter.writerow([dimension[:len(dimension)-1], line1.strip(), getMetadataSection(dimension), "Missing"])
						writtentoFile = True
				elif len(aline1) == 3:
					hier1 = aline1[0] + ";" + aline1[1]
					aggweight1 = aline1[2]
					for line2 in to_file:
						aline2 = line2.split(";")
						if len(aline1) == len(aline2):
							hier2 = aline2[0] + ";" + aline2[1]
							aggweight2 = aline2[2]
							if hier1 == hier2:
								exists = True
								if processfile21 == True:
									writtentoFile = True
								if processfile12 == True:
									if aggweight1 == aggweight2:
										pass
									else:
										CSVWriter.writerow([dimension[:len(dimension)-1], hier2.strip(), 'aggrweight', aggweight1.strip(), aggweight2.strip()])
										writtentoFile = True
					if exists == False and processfile12 == True:
						CSVWriter.writerow([dimension[:len(dimension)-1], line1.strip(), getMetadataSection(dimension), "", "Missing"])
						writtentoFile = True
					elif exists == False and processfile21 == True:
						CSVWriter.writerow([dimension[:len(dimension)-1], line1.strip(), getMetadataSection(dimension), "Missing"])
						writtentoFile = True
				elif processfile12:
					for line2 in to_file:
						aline2 = line2.split(";")
						defaultParentIndex = len(aline2) - 2
						descriptionIndex = len(aline2) - 1
						if aline1[0].upper() == aline2[0].upper() and len(aline1) == len(aline2):
							exists = True
							
							propertyIndex = 0
							for item in aline1:
								if item.upper() != aline2[propertyIndex].upper():
									if (defaultParentIndex == propertyIndex):
										if (item.split("=")[1] == "#root" and aline2[propertyIndex].split("=")[1] =="") or item.split("=")[1] == "" and aline2([propertIndex].split("=")[1] =="#root"):
											pass
										else:
											CSVWriter.writerow([dimension[:len(dimension)-1], aline1[0], aPropertyArray[propertyIndex],item.split("=")[1].strip(), aline2[propertyIndex].split("=")[1].strip()])
									elif (descriptionIndex == propertyIndex) and (item.split("=")[1].strip() ==aline2[propertyIndex].split("=")[1].strip()):
										pass
									else:
										CSVWriter.writerow([dimension[:len(dimension)-1], aline1[0], aPropertyArray[propertyIndex],item.strip(), aline2[propertyIndex].strip()])
								propertyIndex = propertyIndex + 1
				elif processfile21:
					for line2 in to_file:
						aline2 = line2.split(";")
						if aline1[0].upper() == aline2[0].upper() and len(aline1) == len(aline2):
							exists = True
				if exists == False and writtentoFile == False and processfile12 == True:
					CSVWriter.writerow([dimension[:len(dimension)-1], aline1[0].strip(), getMetadataSection(dimension), "", "Missing"])
				elif exists == False and writtentoFile == False and processfile21 == True:
					CSVWriter.writerow([dimension[:len(dimension)-1], aline1[0].strip(), getMetadataSection(dimension), "Missing"])
				i = i + 1
			from_file.close()
			to_file.close()
	

def process():
	startTime = time.time()
	#printLine("Validating the data files, the process will abort if the files are not symmetric")
	parser = argparse.ArgumentParser(formatter_class=argparse.RawDescriptionHelpFormatter, description=textwrap.dedent('''
				This utility is meant for comparing two HFM metadata files in .app or .txt format
				Two app files must be provided for comparison to run'''))
	parser._action_groups.pop()
	required = parser.add_argument_group('Required Arguments - Need to provide them')
	optional = parser.add_argument_group('Optional Arguments')
	required.add_argument('-f1', '--file1', help='Provide the name of the 1st metadata file. default is ABTPLNQA_Metadata.app'\
	, default = 'ABTPLNQA_Metadata.app')
	required.add_argument('-f2', '--file2', help='Provide the name of the 2nd metadata file. default is ABTPROD_Metadata.app'\
    , default = 'ABTPROD_Metadata.app')
	optional.add_argument('-p', '--path', help='Provide the path to the metadata files, if path is not provided\
	then this program will try to check for the files in the current working directory', default ='')
	
	args = vars(parser.parse_args())
	global listOfCustDimensions
	global differencesfilePath 
	
	if args["path"] == "":
		args["path"] = Path.cwd()
	differencesfilePath = getFullPath(args["path"],"Results_" + now.strftime("%Y%m%d_%H%M%S") + ".csv")
	
	if args["file1"] == "" or args["file2"] == "" :
		raise Exception("Terminating the program as Two metadata files are not provided")
	#listOfCustDimensions = numberOfCustomDimensions(args["path"], [args["file1"], args["file2"]])
	processFolder = createTempFolders(args["path"], args["file1"].split(".")[0], args["file2"].split(".")[0])
	trimMetadataFiles(args["path"], [args["file1"], args["file2"]])
	aDimensionsInFiles = splitFile(args["path"], [args["file1"], args["file2"]])
	
	with open(differencesfilePath, "a", encoding="cp1252", newline="") as outputCSVFile:
		CSVWriter = csv.writer(outputCSVFile)
		CSVWriter.writerow(["Dimension", "Member Name", "Property", args["file1"].split(".")[0], args["file2"].split(".")[0]])
	count = 0
	for eachDimensionsFile in aDimensionsInFiles:
		count = count + 1
		for index,fileInMetadatOne in enumerate(eachDimensionsFile):
			
			if count == 1:
				if fileInMetadatOne in aDimensionsInFiles[1]:
					findDifferences(getFullPath(args["path"],"Dimension_files"), args["file1"].split(".")[0], args["file2"].split(".")[0], fileInMetadatOne + ".txt")
				else:
					printWarningLines(processBlock(fileInMetadatOne) + " doesn't exists in " + args["file2"] + "... therefore skipping the comparison")
			else:
				if not fileInMetadatOne in aDimensionsInFiles[0]:
					printWarningLines(processBlock(fileInMetadatOne) + " doesn't exists in " + args["file1"] + "... therefore skipping the comparison")
                
	createResultsFile(differencesfilePath)
	printLines("Deleting the temporary files and directories")
	shutil.rmtree(getFullPath(args["path"], "Dimension_files"), ignore_errors=False, onerror=None)
	os.remove(getFullPath(args["path"],args["file1"].split(".")[0]+"_v1.txt"))
	os.remove(getFullPath(args["path"],args["file2"].split(".")[0]+"_v1.txt"))	
	os.remove(differencesfilePath)
	#writeToResultsFile()
	print("INFO: {}: Results are in the file: {}" . format(getCurrentTime(), excelFilePath))
	#endTime = time.time()
	print("INFO: {}: processing completed in: {} secs" . format(getCurrentTime(), str(round(time.time() - startTime, 2))))
	
if __name__ == "__main__":
	process()
						
			
		
	
    
    